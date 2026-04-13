"""
Excel格式设置通用模块
为operation-ZY所有工具提供统一的Excel格式设置
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def apply_excel_formatting(file_path, logger=None):
    """
    应用统一的Excel格式设置

    格式规则：
    - 字体：微软雅黑 10号
    - 列宽：12
    - 表头：加粗、浅灰底色(#D9D9D9)
    - 数值列：标题和内容右对齐
    - 文字列：标题和内容左对齐
    - 冻结首行
    - 自动筛选

    参数:
        file_path: Excel文件路径
        logger: ValidationLogger实例（可选）
    """
    if logger:
        print("开始应用Excel格式...")

    try:
        wb = load_workbook(file_path)
        ws = wb.active

        # 设置列宽为 12
        for col_idx, col in enumerate(ws.columns, 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 12

        # 判断每列是数值列还是文字列
        # 通过检查前10行数据，如果大多数是数值则认为是数值列
        numeric_columns = set()
        for col_idx, col in enumerate(ws.columns, 1):
            numeric_count = 0
            total_count = 0
            for cell in list(col)[:10]:  # 检查前10行
                if cell.value is not None:
                    total_count += 1
                    if isinstance(cell.value, (int, float)):
                        numeric_count += 1
            # 如果超过一半是数值，则认为是数值列
            if total_count > 0 and numeric_count / total_count > 0.5:
                numeric_columns.add(col_idx)

        # 设置表头样式：微软雅黑、加粗、浅灰底色
        header_fill = PatternFill("solid", fgColor="D9D9D9")
        for col_idx, cell in enumerate(ws[1], 1):
            cell.font = Font(name="Microsoft YaHei", size=10, bold=True)
            cell.fill = header_fill
            # 表头对齐：数值列右对齐，文字列左对齐
            if col_idx in numeric_columns:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # 设置正文样式：微软雅黑、两位小数
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Microsoft YaHei", size=10)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'
                # 正文对齐：数值列右对齐，文字列左对齐
                col_idx = cell.column
                if col_idx in numeric_columns:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # 冻结首行
        ws.freeze_panes = "A2"

        # 添加自动筛选
        max_col = ws.max_column
        max_row = ws.max_row
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

        wb.save(file_path)
        print("Excel格式应用完成")

    except Exception as e:
        print(f"应用Excel格式失败: {e}")
